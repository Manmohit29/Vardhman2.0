from openpyxl import load_workbook
from openpyxl.chart import BarChart, BarChart3D, Reference, Series
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
import smtplib
from os.path import basename
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.utils import formatdate
from datetime import datetime, timedelta
import os
import logging
from config import machine_info

MACHINE = machine_info["name"]
log = logging.getLogger()


def find_last_filled_row(ws):
    return max((row for row in range(ws.max_row, 0, -1)
                if any(cell.value is not None for cell in ws[row])),
               default=None)


def apply_styles(ws, cell_coord, value=None, size=None):
    cell = ws[cell_coord]
    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.fill = PatternFill(start_color="FF5cd65c", end_color="FF5cd65c", fill_type="solid")
    if value:
        cell.value = value
    if size:
        cell.font = Font(size=size)


def delete_rows_from_start(ws, start_row):
    """
    Deletes rows from the given start_row to the last row in the worksheet,
    and unmerges any merged cells in that range.

    Parameters:
    ws (Worksheet): The worksheet object from which rows need to be deleted.
    start_row (int): The starting row number from which to begin deletion.
    """
    total_rows = ws.max_row
    num_rows_to_delete = total_rows - start_row + 1

    merged_ranges_to_unmerge = []
    for merged_cell in ws.merged_cells.ranges:
        min_row, max_row = merged_cell.min_row, merged_cell.max_row
        if min_row >= start_row:
            merged_ranges_to_unmerge.append(merged_cell)

    for merged_cell in merged_ranges_to_unmerge:
        ws.unmerge_cells(str(merged_cell))

    for _ in range(num_rows_to_delete):
        ws.delete_rows(start_row)


def generate_report(today, c):
    global last_filled_data_row
    wb = load_workbook("Machine-template.xlsx")
    ws = wb["Machine Production report"]
    ws['B1'] = f"AT1-{MACHINE} Shift Report"

    delete_rows_from_start(ws, 3)
    data_run = []
    for shift in ['A', 'B', 'C']:
        data_run = data_run + c.get_run_report(today, shift)
    try:
        for i, data_run_p in enumerate(data_run):
            try:
                row = i + 3
                print(i, data_run_p)
                ws[f'B{row}'] = i + 1
                ws[f'C{row}'] = data_run_p[0]
                ws[f'D{row}'] = data_run_p[1]
                ws[f'E{row}'] = data_run_p[2]
                ws[f'F{row}'] = data_run_p[3]
                ws[f'G{row}'] = data_run_p[4]
                ws[f'H{row}'] = data_run_p[5]
                ws[f'I{row}'] = data_run_p[6]
                ws[f'J{row}'] = data_run_p[7]
                ws[f'K{row}'] = round(data_run_p[8] / 60)
                ws[f'L{row}'] = int(data_run_p[9] / (data_run_p[8] / 60))
                ws[f'M{row}'] = round(data_run_p[9])
                ws[f'R{row}'] = round(data_run_p[10], 2)
                ws[f'O{row}'] = round(data_run_p[11], 2)
                ws[f'Q{row}'] = data_run_p[12]
                ws[f'P{row}'] = round(data_run_p[13], 2)
                ws[f'V{row}'] = data_run_p[14]
                ws[f'N{row}'] = round((data_run_p[9] * (data_run_p[3] / 1000)), 2)
                ws[f'S{row}'] = round((data_run_p[11]) / (data_run_p[9] * (data_run_p[3] / 1000)), 2)
                ws[f'T{row}'] = round(data_run_p[13] / (data_run_p[9] * (data_run_p[3] / 1000)), 2)
                ws[f'U{row}'] = round(data_run_p[12] / (data_run_p[9] * (data_run_p[3] / 1000)), 2)

            except Exception as e:
                print(e)

        # Apply border and alignment to all cells within the range B3 to V(last row)
        for row in ws.iter_rows(min_row=3, min_col=2, max_col=22):
            for cell in row:
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                     top=Side(style='thin'), bottom=Side(style='thin'))
                cell.alignment = Alignment(horizontal='center', vertical='center')

        last_filled_row = find_last_filled_row(ws)
        target_row = last_filled_row + 2 if last_filled_row is not None else 1
        ws.merge_cells(start_row=target_row, start_column=2, end_row=target_row, end_column=4)
        ws[f"B{target_row}"].value = "Shift Production"

        columns_to_sum = ['E', 'F', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U']
        for col in columns_to_sum:
            cell = ws[f"{col}{target_row}"]
            cell.value = f"=SUM({col}3:{col}{last_filled_row})"

        for col in range(2, 23):
            cell = ws.cell(row=target_row, column=col)
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                 top=Side(style='thin'), bottom=Side(style='thin'))
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color="FF5cd65c", end_color="FF5cd65c", fill_type="solid")

        last_filled_changeover_data = find_last_filled_row(ws)
        target_row_changeover_data = last_filled_changeover_data + 2 if last_filled_changeover_data is not None else 1
        ws.merge_cells(start_row=target_row_changeover_data, start_column=3,
                       end_row=target_row_changeover_data + 1, end_column=3)
        data_changeover = c.get_changeover_report(today)
        apply_styles(ws, f"C{target_row_changeover_data}", "Avg Changeover", size=12)
        apply_styles(ws, f"C{target_row_changeover_data + 1}")
        apply_styles(ws, f"D{target_row_changeover_data}", "In nos", size=12)
        apply_styles(ws, f"D{target_row_changeover_data + 1}", data_changeover[1], size=12)
        apply_styles(ws, f"E{target_row_changeover_data}")
        apply_styles(ws, f"E{target_row_changeover_data + 1}")
        apply_styles(ws, f"F{target_row_changeover_data}", "In time", size=12)
        apply_styles(ws, f"F{target_row_changeover_data + 1}", data_changeover[0], size=12)



    except Exception as e:
        print(e)

    try:
        ws.title = f"{MACHINE} Production report"
        #### MAKE CHART FOR HEAT
        values1 = Reference(ws, min_row=3, max_row=last_filled_row, min_col=19, max_col=19)
        chart1 = BarChart()
        chart1.title = "Heat consumption / Kg of fabric"
        chart1.y_axis.title = "Kg of fabric"
        chart1.x_axis.title = "Heat consumption"
        chart1.add_data(values1, titles_from_data=True)
        ws.add_chart(chart1, f"H{last_filled_row + 5}")

        ### MAKE CHART FOR KWH
        values2 = Reference(ws, min_row=3, max_row=last_filled_row, min_col=20, max_col=20)
        chart2 = BarChart()
        chart2.title = "kwh/ Kg of fabric"
        chart2.y_axis.title = "Kg of fabric"
        chart2.x_axis.title = "KWH"
        chart2.add_data(values2, titles_from_data=True)
        ws.add_chart(chart2, f"Q{last_filled_row + 5}")

        #### MAKE CHART FOR WATER
        values3 = Reference(ws, min_row=3, max_row=last_filled_row, min_col=21, max_col=21)
        chart3 = BarChart()
        chart3.title = "Water consumption / Kg of fabric"
        chart3.y_axis.title = "Kg of fabric"
        chart3.x_axis.title = "Water consumption"
        chart3.add_data(values3, titles_from_data=True)
        ws.add_chart(chart3, f"N{last_filled_row + 5}")


    except Exception as e:
        print(e)

    ws = wb["Machine Downtime report"]
    ws['B1'] = f"AT1-{MACHINE} downtime Report"
    delete_rows_from_start(ws, 3)
    data_stop = []
    for shift in ['A', 'B', 'C']:
        data_stop = data_stop + c.get_stop_report(today, shift)

    for i, data_stop_p in enumerate(data_stop):
        try:
            row = i + 3
            ws[f'B{row}'] = i + 1
            ws[f'C{row}'] = data_stop_p[0]
            ws[f'D{row}'] = data_stop_p[1]
            ws[f'E{row}'] = data_stop_p[2]
            ws[f'F{row}'] = data_stop_p[3]
            ws[f'G{row}'] = data_stop_p[4]
            ws[f'H{row}'] = data_stop_p[5]
            ws[f'I{row}'] = data_stop_p[6]
            ws[f'J{row}'] = round(data_stop_p[7] / 60)
            ws[f'K{row}'] = ws[f'J{row}'].value / 1440  # Calculate K as J/1440
            ws[f'K{row}'].number_format = '0.00%'  # Format K as percentage
            ws[f'L{row}'] = round(data_stop_p[8], 2)
            ws[f'M{row}'] = round(data_stop_p[9], 2)
            ws[f'N{row}'] = int(data_stop_p[10])
            ws[f'O{row}'] = round(data_stop_p[11], 2)
            ws[f'P{row}'] = data_stop_p[12]

        except Exception as e:
            print(e)

        for row in ws.iter_rows(min_row=3, min_col=2, max_col=16):
            for cell in row:
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                     top=Side(style='thin'), bottom=Side(style='thin'))
                cell.alignment = Alignment(horizontal='center', vertical='center')

    last_row = len(data_stop) + 3

    ws[f'I{last_row}'] = "Downtime"
    ws.merge_cells(start_row=last_row, start_column=2, end_row=last_row, end_column=8)
    for col in range(ord('J'), ord('O') + 1):
        col_letter = chr(col)
        ws[f'{col_letter}{last_row}'] = f'=SUM({col_letter}3:{col_letter}{last_row - 1})'

        if col_letter == 'K':
            ws[f'{col_letter}{last_row}'].number_format = '0.00%'

    for col in range(ord('B'), ord('O') + 1):
        col_letter = chr(col)
        cell = ws[f'{col_letter}{last_row}']
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color="FF5cd65c", end_color="FF5cd65c", fill_type="solid")

    border_style = Border(left=Side(style='thin'), right=Side(style='thin'),
                          top=Side(style='thin'), bottom=Side(style='thin'))

    utilization_row = last_row + 1
    ws[f'I{utilization_row}'] = "Utilization"
    ws[f'J{utilization_row}'] = f'=1440 - J{last_row}'
    ws[f'K{utilization_row}'] = f'=ROUND(J{utilization_row} / 1440, 2)'
    fill_style = PatternFill(start_color="FF5cd65c", end_color="FF5cd65c", fill_type="solid")

    # Apply fill and border to Utilization row
    for cell_ref in [f'I{utilization_row}', f'J{utilization_row}', f'K{utilization_row}']:
        ws[cell_ref].fill = fill_style
        ws[cell_ref].border = border_style
        ws[cell_ref].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells(start_row=utilization_row, start_column=2, end_row=utilization_row, end_column=8)

    mr_row = utilization_row + 1
    ws[f'C{mr_row}'] = "MTTR"
    ws[f'D{mr_row}'] = "Repairing time/ Nos of breakdown"
    ws[f'H{mr_row}'] = c.get_MTTR_report(today)
    ws[f'C{mr_row}'].fill = fill_style

    # Apply fill and border to MTTR row
    for cell_ref in [f'C{mr_row}', f'D{mr_row}', f'H{mr_row}']:
        ws[cell_ref].fill = fill_style
        ws[cell_ref].border = border_style
        ws[cell_ref].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells(start_row=mr_row, start_column=4, end_row=mr_row, end_column=7)

    mf_row = mr_row + 1
    ws[f'C{mf_row}'] = "MTBF"
    ws[f'D{mf_row}'] = "Total production time /Nos of breakdown"
    ws[f'H{mf_row}'] = c.get_MTBF_report(today)
    ws[f'C{mf_row}'].fill = fill_style

    # Apply fill and border to MTBF row
    for cell_ref in [f'C{mf_row}', f'D{mf_row}', f'H{mf_row}']:
        ws[cell_ref].fill = fill_style
        ws[cell_ref].border = border_style
        ws[cell_ref].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells(start_row=mf_row, start_column=4, end_row=mf_row, end_column=7)

    try:
        ws.title = f"{MACHINE} Downtime report"
    except Exception as e:
        print(e)

    ws = wb["Downtime Chart"]
    data = c.get_downtown_report(today)
    for i, data_p in enumerate(data):
        row = i + 3
        ws[f'B{row}'] = data_p[0]
        ws[f'C{row}'] = data_p[1]
        ws[f'B{row}'].border = Border(left=Side(border_style="thin"),
                                      right=Side(border_style="thin"),
                                      top=Side(border_style="thin"),
                                      bottom=Side(border_style="thin"))

        ws[f'C{row}'].border = Border(left=Side(border_style="thin"),
                                      right=Side(border_style="thin"),
                                      top=Side(border_style="thin"),
                                      bottom=Side(border_style="thin"))

    ws = wb["Monthly Utilization Trend"]
    ws['C2'] = f"{MACHINE} Utilization & Production Day by Day"

    curr_month = datetime.strptime(today, "%Y-%m-%d")  # .strftime("%Y-%m")
    print("curr_month", curr_month)
    curr_day = int(datetime.strptime(today, "%Y-%m-%d").strftime("%d"))
    print("CD", curr_day)
    ws['B1'].value = curr_month.strftime("%B-%Y")
    for i in range(1, curr_day + 1):
        date_uti_dt = (curr_month - timedelta(days=curr_day) + timedelta(days=i))
        date_uti = date_uti_dt.strftime("%Y-%m-%d")
        date_uti_f = date_uti_dt.strftime("%d-%m-%Y")
        print("date_uti", date_uti)
        print(date_uti)
        utilization, production, lot_number, no_program, breakdown, total_energy, total_fluid, \
        total_air = c.get_utilization_report(date_uti)
        row = i + 3
        ws[f'B{row}'] = date_uti_f
        # ws[f'C{row}'] = utilization
        ws[f'C{row}'] = min(100, max(0, utilization))
        ws[f'D{row}'] = round(production, 2)
        ws[f'E{row}'] = lot_number
        ws[f'F{row}'] = int(no_program)
        ws[f'G{row}'] = breakdown
        ws[f'H{row}'] = (total_energy)
        ws[f'J{row}'] = round(total_fluid, 2)
        ws[f'K{row}'] = round(total_air, 2)

    log.info(data)

    ws = wb["PO wise Prod Summary"]

    data_run_sum = c.get_po_prod_summary(today)
    total_meters = 0
    total_duration = 0
    total_fluid = 0
    total_energy = 0

    try:
        for i, data_run_p in enumerate(data_run_sum):
            try:
                row = i + 4
                print(i, data_run_p)
                if data_run_p[0] == '':
                    ws[f'A{row}'] = "Blank"
                else:
                    ws[f'A{row}'] = data_run_p[0]
                ws[f'B{row}'] = round(data_run_p[1])
                ws[f'C{row}'] = round(data_run_p[2] / 60)
                ws[f'D{row}'] = data_run_p[3]
                ws[f'E{row}'] = data_run_p[4]

                total_meters += round(data_run_p[1])
                total_duration += round(data_run_p[2] / 60)
                total_fluid += data_run_p[3]
                total_energy += data_run_p[4]

                for col in ['A', 'B', 'C', 'D', 'E']:
                    ws[f'{col}{row}'].border = Border(
                        left=Side(border_style='thin'),
                        right=Side(border_style='thin'),
                        top=Side(border_style='thin'),
                        bottom=Side(border_style='thin')
                    )
                    ws[f'{col}{row}'].alignment = Alignment(horizontal='center', vertical='center')
            except Exception as e:
                print(e)

            total_row = len(data_run_sum) + 4
            ws[f'A{total_row}'] = "Grand Total"
            ws[f'B{total_row}'] = total_meters
            ws[f'C{total_row}'] = total_duration
            ws[f'D{total_row}'] = total_fluid
            ws[f'E{total_row}'] = total_energy

            sky_blue_fill = PatternFill(start_color='87CEEB', end_color='87CEEB', fill_type='solid')
            bold_font = Font(bold=True)

            for col in ['A', 'B', 'C', 'D', 'E']:
                ws[f'{col}{total_row}'].border = Border(
                    left=Side(border_style='thin'),
                    right=Side(border_style='thin'),
                    top=Side(border_style='thin'),
                    bottom=Side(border_style='thin')
                )
                ws[f'{col}{total_row}'].alignment = Alignment(horizontal='center', vertical='center')
                ws[f'{col}{total_row}'].fill = sky_blue_fill
                ws[f'{col}{total_row}'].font = bold_font


    except Exception as e:
        print(e)

    ws = wb["CO Downtime Summary"]

    data_stop_sum = c.get_co_downtime_summary(today)

    total_duration = 0
    total_fluid = 0
    total_energy = 0

    try:
        for i, data_run_p in enumerate(data_stop_sum):
            try:
                row = i + 4
                print(i, data_run_p)
                if data_run_p[0] == '':
                    ws[f'A{row}'] = "Blank"
                else:
                    ws[f'A{row}'] = data_run_p[0]
                ws[f'B{row}'] = round(data_run_p[1] / 60)
                ws[f'C{row}'] = data_run_p[2]
                ws[f'D{row}'] = data_run_p[3]

                total_duration += round(data_run_p[1] / 60)
                total_fluid += data_run_p[2]
                total_energy += data_run_p[3]

                for col in ['A', 'B', 'C', 'D']:
                    ws[f'{col}{row}'].border = Border(
                        left=Side(border_style='thin'),
                        right=Side(border_style='thin'),
                        top=Side(border_style='thin'),
                        bottom=Side(border_style='thin')
                    )
                    ws[f'{col}{row}'].alignment = Alignment(horizontal='center', vertical='center')

            except Exception as e:
                print(e)

        total_row = len(data_stop_sum) + 4
        ws[f'A{total_row}'] = "Grand Total"
        ws[f'B{total_row}'] = total_duration
        ws[f'C{total_row}'] = total_fluid
        ws[f'D{total_row}'] = total_energy

        sky_blue_fill = PatternFill(start_color='87CEEB', end_color='87CEEB', fill_type='solid')
        bold_font = Font(bold=True)

        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{total_row}'].border = Border(
                left=Side(border_style='thin'),
                right=Side(border_style='thin'),
                top=Side(border_style='thin'),
                bottom=Side(border_style='thin')
            )
            ws[f'{col}{total_row}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'{col}{total_row}'].fill = sky_blue_fill
            ws[f'{col}{total_row}'].font = bold_font

    except Exception as e:
        print(e)

    file_name = f"AT1-{MACHINE} Mc Report-{today}.xlsx"
    path_ = f"Reports/{file_name}"
    wb.save(path_)
    return path_, file_name


def send_mail(send_from, send_to, subject, text, files=None, server="172.28.0.254", port=25,
              username='at1iiot@vardhman.com', password='',
              use_tls=True):
    """Compose and send email with provided info and attachments.

    Args:
        send_from (str): from name 'Grinding Machine<at1iiot@vardhman.com>'
        send_to (list[str]): to name(s) at1iiot@vardhman.com
        subject (str): message title
        message (str): message body
        files (list[str]): list of file paths to be attached to email
        server (str): mail server host name
        port (int): port number
        username (str): server auth username
        password (str): server auth password
        use_tls (bool): use TLS mode
    """
    assert isinstance(send_to, list)
    msg = MIMEMultipart()
    msg['From'] = send_from
    msg['To'] = ",".join(send_to)
    msg['Date'] = formatdate(localtime=True)
    msg['Subject'] = subject

    msg.attach(MIMEText(text))

    for f in files or []:
        with open(f, "rb") as fil:
            part = MIMEApplication(
                fil.read(),
                Name=basename(f)
            )
        # After the file is closed
        part['Content-Disposition'] = 'attachment; filename="%s"' % basename(f)
        msg.attach(part)

    smtp = smtplib.SMTP(server, port, timeout=10)
    smtp.set_debuglevel(0)
    smtp.connect(server, port)
    # smtp.helo()
    if use_tls:
        smtp.starttls()
    # smtp.ehlo()
    # smtp.login(username, password)
    smtp.sendmail(send_from, send_to, msg.as_string())
    smtp.close()
